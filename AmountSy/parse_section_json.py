"""
段落JSON解析脚本
将Full_Section.json解析为多个CSV表格
"""

import json
import pandas as pd
from pathlib import Path
from collections import Counter


def load_json(file_path):
    """加载JSON文件"""
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def parse_section_summary(data):
    """解析段落汇总信息"""
    sections = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        section_list = scene_data.get('sections', [])

        for section in section_list:
            section_name = section.get('section_name', '')
            tracks = section.get('tracks', [])
            animations = section.get('animset_animations', {})

            # 统计各种轨道类型的数量
            track_types = [track.get('track_type') for track in tracks]
            track_type_count = Counter(track_types)

            # 统计剪辑总数
            total_clips = sum(len(track.get('clips', [])) for track in tracks)

            sections.append({
                '场景路径': scene_path,
                '段落名称': section_name,
                '轨道总数': len(tracks),
                '类型1轨道数': track_type_count.get(1, 0),
                '类型2轨道数': track_type_count.get(2, 0),
                '类型3轨道数': track_type_count.get(3, 0),
                '类型4轨道数': track_type_count.get(4, 0),
                '类型5轨道数': track_type_count.get(5, 0),
                '类型6轨道数': track_type_count.get(6, 0),
                '类型7轨道数': track_type_count.get(7, 0),
                '类型8轨道数': track_type_count.get(8, 0),
                'Clip总数': total_clips,
                '动画种类数': len(animations),
                '动画使用总次数': sum(animations.values()) if animations else 0
            })

    return pd.DataFrame(sections)


def parse_tracks(data):
    """解析轨道详情"""
    track_records = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        section_list = scene_data.get('sections', [])

        for section in section_list:
            section_name = section.get('section_name', '')
            tracks = section.get('tracks', [])

            for idx, track in enumerate(tracks, 1):
                track_name = track.get('track_name', '')
                track_type = track.get('track_type', '')
                clips = track.get('clips', [])

                track_records.append({
                    '场景路径': scene_path,
                    'Section名称': section_name,
                    '轨道序号': idx,
                    '轨道名称': track_name,
                    '轨道类型': track_type,
                    '剪辑数量': len(clips)
                })

    return pd.DataFrame(track_records)


def parse_clips(data):
    """解析剪辑详情"""
    clip_records = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        section_list = scene_data.get('sections', [])

        for section in section_list:
            section_name = section.get('section_name', '')
            tracks = section.get('tracks', [])

            for track in tracks:
                track_name = track.get('track_name', '')
                track_type = track.get('track_type', '')
                clips = track.get('clips', [])

                for clip_idx, clip_type in enumerate(clips, 1):
                    clip_records.append({
                        '场景路径': scene_path,
                        '段落名称': section_name,
                        '轨道名称': track_name,
                        '轨道类型': track_type,
                        '剪辑序号': clip_idx,
                        '剪辑类型': clip_type
                    })

    return pd.DataFrame(clip_records)


def parse_section_animations(data):
    """解析段落动画信息"""
    animation_records = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        section_list = scene_data.get('sections', [])

        for section in section_list:
            section_name = section.get('section_name', '')
            animations = section.get('animset_animations', {})

            for anim_name, count in animations.items():
                animation_records.append({
                    '场景路径': scene_path,
                    '段落名称': section_name,
                    '动画名称': anim_name,
                    '使用次数': count
                })

    return pd.DataFrame(animation_records)


def parse_track_type_statistics(data):
    """解析轨道类型统计"""
    type_records = []

    scene_solutions = data.get('scene_solutions', {})

    # 收集所有轨道类型和名称的对应关系
    track_type_names = {}

    for scene_path, scene_data in scene_solutions.items():
        section_list = scene_data.get('sections', [])

        for section in section_list:
            tracks = section.get('tracks', [])

            for track in tracks:
                track_type = track.get('track_type', '')
                track_name = track.get('track_name', '')

                if track_type not in track_type_names:
                    track_type_names[track_type] = set()
                track_type_names[track_type].add(track_name)

    # 生成统计记录
    for track_type, names in sorted(track_type_names.items()):
        type_records.append({
            '轨道类型': track_type,
            '常见轨道名称': ', '.join(sorted(names)[:10]),  # 只显示前10个
            '不同名称数': len(names)
        })

    return pd.DataFrame(type_records)


def parse_clip_type_statistics(data):
    """解析剪辑类型统计"""
    clip_counter = Counter()

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        section_list = scene_data.get('sections', [])

        for section in section_list:
            tracks = section.get('tracks', [])

            for track in tracks:
                clips = track.get('clips', [])
                clip_counter.update(clips)

    # 生成统计记录
    clip_records = [
        {'剪辑类型': clip_type, '出现次数': count}
        for clip_type, count in clip_counter.most_common()
    ]

    return pd.DataFrame(clip_records)


def main():
    """主函数"""
    # 文件路径
    json_file = 'Full_Section.json'
    output_dir = Path('.')

    print(f"正在读取JSON文件: {json_file}")
    data = load_json(json_file)

    # 解析各个表格
    print("正在解析段落汇总信息...")
    df_section_summary = parse_section_summary(data)

    print("正在解析轨道详情...")
    df_tracks = parse_tracks(data)

    print("正在解析剪辑详情...")
    df_clips = parse_clips(data)

    print("正在解析段落动画信息...")
    df_animations = parse_section_animations(data)

    print("正在解析轨道类型统计...")
    df_track_types = parse_track_type_statistics(data)

    print("正在解析剪辑类型统计...")
    df_clip_types = parse_clip_type_statistics(data)

    # 保存为CSV文件
    print("\n正在保存CSV文件...")
    df_section_summary.to_csv(output_dir / 'section_summary.csv', index=False, encoding='utf-8-sig')
    print(f"  - section_summary.csv (段落汇总: {len(df_section_summary)} 行)")

    df_tracks.to_csv(output_dir / 'section_tracks.csv', index=False, encoding='utf-8-sig')
    print(f"  - section_tracks.csv (轨道详情: {len(df_tracks)} 行)")

    df_clips.to_csv(output_dir / 'section_clips.csv', index=False, encoding='utf-8-sig')
    print(f"  - section_clips.csv (剪辑详情: {len(df_clips)} 行)")

    df_animations.to_csv(output_dir / 'section_animations.csv', index=False, encoding='utf-8-sig')
    print(f"  - section_animations.csv (段落动画: {len(df_animations)} 行)")

    df_track_types.to_csv(output_dir / 'track_type_statistics.csv', index=False, encoding='utf-8-sig')
    print(f"  - track_type_statistics.csv (轨道类型统计: {len(df_track_types)} 行)")

    df_clip_types.to_csv(output_dir / 'clip_type_statistics.csv', index=False, encoding='utf-8-sig')
    print(f"  - clip_type_statistics.csv (剪辑类型统计: {len(df_clip_types)} 行)")

    # 保存为单个Excel文件的多个工作表
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        excel_file = output_dir / 'section_analysis.xlsx'
        print(f"\n正在保存Excel文件: {excel_file}")
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 写入各个工作表
            sheets_data = [
                (df_section_summary, '段落汇总'),
                (df_tracks, '轨道详情'),
                (df_clips, '剪辑详情'),
                (df_animations, '段落动画'),
                (df_track_types, '轨道类型统计'),
                (df_clip_types, '剪辑类型统计')
            ]

            for df, sheet_name in sheets_data:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # 获取工作表对象
                worksheet = writer.sheets[sheet_name]

                # 设置表头样式
                header_font = Font(bold=True, color='FFFFFF', size=11)
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')

                # 应用表头样式
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # 设置行高
                worksheet.row_dimensions[1].height = 25  # 表头行高
                for row in range(2, worksheet.max_row + 1):
                    worksheet.row_dimensions[row].height = 20  # 数据行高

                # 自动调整列宽
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass

                    # 设置列宽，最小12，最大60
                    adjusted_width = min(max(max_length + 2, 12), 60)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # 冻结首行
                worksheet.freeze_panes = 'A2'

                # 添加自动筛选
                worksheet.auto_filter.ref = worksheet.dimensions

                # 设置数据单元格对齐方式
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(vertical='center', wrap_text=False)

        print(f"  Excel文件保存成功! (已优化格式)")
    except ImportError:
        print("  提示: 需要安装openpyxl才能保存Excel文件")
        print("  运行: pip install openpyxl")

    print("\n解析完成!")
    print(f"\n数据统计:")
    print(f"  - 段落总数: {len(df_section_summary)}")
    print(f"  - 轨道总数: {len(df_tracks)}")
    print(f"  - 剪辑总数: {len(df_clips)}")
    print(f"  - 动画记录数: {len(df_animations)}")
    print(f"  - 轨道类型数: {len(df_track_types)}")
    print(f"  - 剪辑类型数: {len(df_clip_types)}")


if __name__ == '__main__':
    main()
