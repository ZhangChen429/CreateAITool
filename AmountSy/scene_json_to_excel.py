#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Scene Solution JSON to Excel Converter
将 scene solution 扫描生成的 JSON 文件转换为 Excel 表格
"""

import json
import sys
import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def load_json(json_path):
    """加载JSON文件"""
    print(f"正在读取JSON文件: {json_path}")
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    print(f"✓ JSON文件读取成功")
    return data


def parse_scene_solutions(data):
    """解析scene solutions数据（过滤含version/versions的路径，统计中断场景和事件标签数量）"""
    scene_solutions = data.get('scene_solutions', {})

    results = {
        'summary': [],  # 包含每个场景的关键统计（含数量统计）
        'interruption_scenarios_count': [],  # 按场景分组的中断场景数量
        'event_execution_tags_count': [],  # 按场景分组的事件标签数量
        'actors_and_props_count': [],  # 用于单独统计演员和道具数量
        'node_type_count': [],
        'node_type_count_no_version': []
    }

    for scene_path, scene_data in scene_solutions.items():
        # 过滤逻辑：忽略含 \version\、\versions\、/version/、/versions/ 的路径（不区分大小写）
        lower_path = scene_path.lower()
        if (
                '\\version\\' in lower_path or
                '/version/' in lower_path or
                '\\versions\\' in lower_path or
                '/versions/' in lower_path
        ):
            continue

        # 1. 摘要信息（新增中断场景数量和事件标签数量）
        interruption_count = len(scene_data.get('interruption_scenarios', []))
        event_tags_count = len(scene_data.get('event_execution_tags', []))
        # 2. 提取 actors_count 和 props_count
        actors_count = scene_data.get('actors_count', 0)
        props_count = scene_data.get('props_count', 0)
        # 提取 rid_animations 字典的键值对数量
        rid_animations_count = len(scene_data.get('rid_animations', {}))
        # 提取 animset_animations 字典的键值对数量 (作为额外参考)
        Animal_animations_count = len(scene_data.get('Animal_animations', {}))
        # 提取 reference_points_count
        reference_points_count = scene_data.get('reference_points_count', 0)  # <<<--- 新增提取

        summary_row = {
            'Scene Path': scene_path,
            'Interruption Scenarios Count': interruption_count,  # 数组长度统计
            'Event Execution Tags Count': event_tags_count,  # 数组长度统计
            'Actors Count': actors_count,  # <<<--- 新增
            'Props Count': props_count,  # <<<--- 新增
            'RID Animations Count': rid_animations_count,  # <<<--- 新增 RID 动画数量
            'Reference Points Count': reference_points_count,# <<<--- 新增 Point
            'Animal Animations Count': Animal_animations_count,  # <<<--- 额外添加 Animal 动画数量

            'Total Nodes': scene_data.get('scene_graph_nodes', {}).get('total_nodes', 0)
        }
        results['summary'].append(summary_row)

        # 2. 中断场景数量（单独分组统计，便于后续分析）
        results['interruption_scenarios_count'].append({
            'Scene Path': scene_path,
            'interruption_scenarios_count': interruption_count  # 直接存储长度
        })

        # 3. 事件执行标签数量（单独分组统计）
        results['event_execution_tags_count'].append({
            'Scene Path': scene_path,
            'event_execution_tags_count': event_tags_count  # 直接存储长度
        })

        # 5. 演员和道具数量（新增：单独分组统计，便于后续分析）
        results['actors_and_props_count'].append({
            'Scene Path': scene_path,
            'Actors Count': actors_count,
            'Props Count': props_count
        })

        # 4. 节点类型统计（原始数据）
        node_counts = scene_data.get('scene_graph_nodes', {}).get('node_type_count', {})
        for node_type, count in node_counts.items():
            results['node_type_count'].append({
                'Scene Path': scene_path,
                'Node Type': node_type,
                'Count': count
            })

        # 5. 过滤版本路径后的节点类型统计（因场景已过滤，此处直接复用统计）
        for node_type, count in node_counts.items():
            results['node_type_count_no_version'].append({
                'Scene Path': scene_path,
                'Node Type': node_type,
                'Count': count
            })



    return results


def create_excel(results, output_path):
    """创建Excel文件（修复：Sheet名称去掉非法字符/）"""
    print(f"\n正在创建Excel文件: {output_path}")

    writer = pd.ExcelWriter(output_path, engine='openpyxl')

    # Sheet 1: 摘要
    if results.get('summary', []):
        df_summary = pd.DataFrame(results['summary'])
        # 避免空DataFrame导出
        if not df_summary.empty:
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            print(f"  ✓ Sheet 'Summary' - {len(df_summary)} 条记录")

    # Sheet 2: Interruption Scenarios
    if results.get('interruption_scenarios_count', []):
        df_scenarios = pd.DataFrame(results['interruption_scenarios_count'])
        if not df_scenarios.empty:
            df_scenarios.to_excel(writer, sheet_name='Interruption Scenarios Count', index=False)


    # Sheet 3: Event Execution Tags
    if results.get('event_execution_tags_count', []):
        df_tags = pd.DataFrame(results['event_execution_tags_count'])
        if not df_tags.empty:
            df_tags.to_excel(writer, sheet_name='Event Execution Tags Count', index=False)

    # Sheet 4: Actors and Props Count (新增 Sheet)
    if results.get('actors_and_props_count', []):
            df_actors_props = pd.DataFrame(results['actors_and_props_count'])
            if not df_actors_props.empty:
                df_actors_props.to_excel(writer, sheet_name='Actors and Props Count', index=False)

    # Sheet 4: Node Type Count (详细)
    if results['node_type_count']:
        df_nodes = pd.DataFrame(results['node_type_count'])
        df_nodes.to_excel(writer, sheet_name='Node Type Count (Detail)', index=False)
        print(f"  ✓ Sheet 'Node Type Count (Detail)' - {len(df_nodes)} 条记录")

    # Sheet 5: Node Type Count (汇总)
    if results['node_type_count']:
        df_nodes_summary = pd.DataFrame(results['node_type_count'])
        node_summary = df_nodes_summary.groupby('Node Type')['Count'].sum().reset_index()
        node_summary = node_summary.sort_values('Count', ascending=False)
        node_summary.to_excel(writer, sheet_name='Node Type Count (Summary)', index=False)
        print(f"  ✓ Sheet 'Node Type Count (Summary)' - {len(node_summary)} 种节点类型")
    writer.close()

    # 美化Excel
    beautify_excel(output_path)

    print(f"\n✓ Excel文件创建成功!")


def beautify_excel(excel_path):
    """美化Excel表格"""
    from openpyxl import load_workbook

    wb = load_workbook(excel_path)

    # 定义样式
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 设置表头样式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_path)


def main():
    """主函数"""
    print("=" * 60)
    print("Scene Solution JSON to Excel Converter")
    print("=" * 60)
    print()

    # 获取输入文件
    if len(sys.argv) > 1:
        json_path = sys.argv[1]
    else:
        json_path = r"D:\AppSoft\Sy2077\ResourceWrite\scene_detailed.json"

    if not os.path.exists(json_path):
        print(f"❌ 错误: 文件不存在: {json_path}")
        return

    # 生成输出文件名
    output_path = "Scene_nodes_Section.xlsx"

    # 检查文件是否已存在（修复：之前的判断逻辑错误，现在正确判断文件是否存在）
    if os.path.exists(output_path):
        overwrite = input(f"\n文件已存在: {output_path}\n是否覆盖? (y/n): ").lower()
        if overwrite != 'y':
            print("操作已取消")
            return

    try:
        # 1. 加载JSON
        data = load_json(json_path)

        # 2. 解析数据
        print("\n正在解析数据...")
        results = parse_scene_solutions(data)

        # 3. 创建Excel
        create_excel(results, output_path)

        # 4. 显示统计信息
        print("\n" + "=" * 60)
        print("统计信息:")
        print(f"  总场景数: {len(results['summary'])}")
        print(f"  忽略/version/后的节点记录数: {len(results['node_type_count_no_version'])}")
        print(f"  Interruption Scenarios: {len(results['interruption_scenarios_count'])}")
        print(f"  Event Execution Tags: {len(results['event_execution_tags_count'])}")
        print(f"  节点类型记录数: {len(results['node_type_count'])}")
        print("=" * 60)

    except Exception as e:
        print(f"\n❌ 错误: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    main()
    input("\n按回车键退出...")