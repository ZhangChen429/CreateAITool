"""
场景JSON解析脚本
将Empty_Questscenes.json解析为多个CSV表格
"""

import json
import pandas as pd
from pathlib import Path


def load_json(file_path):
    """加载JSON文件"""
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def parse_scene_summary(data):
    """解析场景汇总信息"""
    scenes = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        # 跳过包含 versions 的场景路径
        if 'versions' in scene_path.lower():
            continue

        # 获取动画数据
        animset_animations = scene_data.get('animset_animations', {})
        playsk_animations = scene_data.get('PlaySkAnim_animations', {})
        changeidle_animations = scene_data.get('ChangeIdleAnim_animations', {})
        changework_animations = scene_data.get('ChangeWorkEvent_animations', {})
        stopwork_animations = scene_data.get('StopWorkEvent_animations', {})

        scene_info = {
            '场景路径': scene_path,
            '中断场景数': len(scene_data.get('interruption_scenarios', [])),
            '事件执行标签数': len(scene_data.get('event_execution_tags', [])),
            '总节点数': scene_data.get('scene_graph_nodes', {}).get('total_nodes', 0),
            '演员数量': scene_data.get('actors_count', 0),
            '道具数量': scene_data.get('props_count', 0),
            '参考点数量': scene_data.get('reference_points_count', 0),
            '外部工作点数': len(scene_data.get('workspots', {}).get('external_workspots', [])),
            '内部工作点数': len(scene_data.get('workspots', {}).get('internal_workspots', [])),
            '动画总数': sum(animset_animations.values()) if animset_animations else 0,
            'PlaySkAnim种类数': len(playsk_animations),
            'PlaySkAnim使用次数': sum(playsk_animations.values()) if playsk_animations else 0,
            'ChangeIdleAnim种类数': len(changeidle_animations),
            'ChangeIdleAnim使用次数': sum(changeidle_animations.values()) if changeidle_animations else 0,
            'ChangeWorkEvent种类数': len(changework_animations),
            'ChangeWorkEvent使用次数': sum(changework_animations.values()) if changework_animations else 0,
            'StopWorkEvent种类数': len(stopwork_animations),
            'StopWorkEvent使用次数': sum(stopwork_animations.values()) if stopwork_animations else 0
        }
        scenes.append(scene_info)

    return pd.DataFrame(scenes)


def parse_node_types(data):
    """解析节点类型统计"""
    node_records = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        # 跳过包含 versions 的场景路径
        if 'versions' in scene_path.lower():
            continue
        node_type_count = scene_data.get('scene_graph_nodes', {}).get('node_type_count', {})

        for node_type, count in node_type_count.items():
            node_records.append({
                '场景路径': scene_path,
                '节点类型': node_type,
                '数量': count
            })

    return pd.DataFrame(node_records)


def parse_workspots(data):
    """解析工作点信息"""
    workspot_records = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        # 跳过包含 versions 的场景路径
        if 'versions' in scene_path.lower():
            continue
        workspots = scene_data.get('workspots', {})

        # 外部工作点
        for workspot in workspots.get('external_workspots', []):
            path = workspot.get('path', '')
            animsets = workspot.get('animsets', [])
            workspot_records.append({
                '场景路径': scene_path,
                '工作点类型': '外部',
                '工作点路径': path,
                'Animsets数量': len(animsets)
            })

        # 内部工作点
        for workspot in workspots.get('internal_workspots', []):
            # 新结构使用 name 字段，旧结构使用 path 字段
            name_or_path = workspot.get('name', workspot.get('path', ''))
            workspot_records.append({
                '场景路径': scene_path,
                '工作点类型': '内部',
                '工作点路径': name_or_path,
                'Animsets数量': 0
            })

    return pd.DataFrame(workspot_records)


def parse_animsets(data):
    """解析Animset信息"""
    animset_records = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        # 跳过包含 versions 的场景路径
        if 'versions' in scene_path.lower():
            continue
        workspots = scene_data.get('workspots', {})

        # 外部工作点的animsets
        for workspot in workspots.get('external_workspots', []):
            workspot_path = workspot.get('path', '')
            animsets = workspot.get('animsets', [])

            for animset_path in animsets:
                animset_records.append({
                    '场景路径': scene_path,
                    '工作点路径': workspot_path,
                    'Animset路径': animset_path
                })

    return pd.DataFrame(animset_records)


def parse_animations(data):
    """解析动画信息"""
    animation_records = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        # 跳过包含 versions 的场景路径
        if 'versions' in scene_path.lower():
            continue
        animations = scene_data.get('animset_animations', {})

        for anim_name, count in animations.items():
            animation_records.append({
                '场景路径': scene_path,
                '动画名称': anim_name,
                '使用次数': count
            })

    return pd.DataFrame(animation_records)


def parse_playsk_animations(data):
    """解析PlaySkAnim动画信息"""
    animation_records = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        # 跳过包含 versions 的场景路径
        if 'versions' in scene_path.lower():
            continue
        animations = scene_data.get('PlaySkAnim_animations', {})

        for anim_name, count in animations.items():
            animation_records.append({
                '场景路径': scene_path,
                '动画名称': anim_name,
                '使用次数': count
            })

    return pd.DataFrame(animation_records)


def parse_changeidle_animations(data):
    """解析ChangeIdleAnim动画信息"""
    animation_records = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        # 跳过包含 versions 的场景路径
        if 'versions' in scene_path.lower():
            continue
        animations = scene_data.get('ChangeIdleAnim_animations', {})

        for anim_name, count in animations.items():
            animation_records.append({
                '场景路径': scene_path,
                '动画名称': anim_name,
                '使用次数': count
            })

    return pd.DataFrame(animation_records)


def parse_changework_animations(data):
    """解析ChangeWorkEvent动画信息"""
    animation_records = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        # 跳过包含 versions 的场景路径
        if 'versions' in scene_path.lower():
            continue
        animations = scene_data.get('ChangeWorkEvent_animations', {})

        for anim_name, count in animations.items():
            animation_records.append({
                '场景路径': scene_path,
                '动画名称': anim_name,
                '使用次数': count
            })

    return pd.DataFrame(animation_records)


def parse_stopwork_animations(data):
    """解析StopWorkEvent动画信息"""
    animation_records = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        # 跳过包含 versions 的场景路径
        if 'versions' in scene_path.lower():
            continue
        animations = scene_data.get('StopWorkEvent_animations', {})

        for anim_name, count in animations.items():
            animation_records.append({
                '场景路径': scene_path,
                '动画名称': anim_name,
                '使用次数': count
            })

    return pd.DataFrame(animation_records)


def parse_interruption_scenarios(data):
    """解析中断场景信息"""
    interruption_records = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        # 跳过包含 versions 的场景路径
        if 'versions' in scene_path.lower():
            continue
        scenarios = scene_data.get('interruption_scenarios', [])

        for idx, scenario in enumerate(scenarios, 1):
            interruption_records.append({
                '场景路径': scene_path,
                '序号': idx,
                '类型': scenario.get('type', '')
            })

    return pd.DataFrame(interruption_records)


def parse_event_execution_tags(data):
    """解析事件执行标签信息"""
    tag_records = []

    scene_solutions = data.get('scene_solutions', {})

    for scene_path, scene_data in scene_solutions.items():
        # 跳过包含 versions 的场景路径
        if 'versions' in scene_path.lower():
            continue
        tags = scene_data.get('event_execution_tags', [])

        for idx, tag in enumerate(tags, 1):
            tag_records.append({
                '场景路径': scene_path,
                '序号': idx,
                '类型': tag.get('type', '')
            })

    return pd.DataFrame(tag_records)


def main():
    """主函数"""
    # 文件路径
    json_file = 'QuestscenesAnimalsq004.json'
    output_dir = Path('.')

    print(f"正在读取JSON文件: {json_file}")
    data = load_json(json_file)

    # 解析各个表格
    print("正在解析场景汇总信息...")
    df_summary = parse_scene_summary(data)

    print("正在解析节点类型统计...")
    df_nodes = parse_node_types(data)

    print("正在解析工作点信息...")
    df_workspots = parse_workspots(data)

    print("正在解析Animset信息...")
    df_animsets = parse_animsets(data)

    print("正在解析动画信息...")
    df_animations = parse_animations(data)

    print("正在解析PlaySkAnim动画...")
    df_playsk_animations = parse_playsk_animations(data)

    print("正在解析ChangeIdleAnim动画...")
    df_changeidle_animations = parse_changeidle_animations(data)

    print("正在解析ChangeWorkEvent动画...")
    df_changework_animations = parse_changework_animations(data)

    print("正在解析StopWorkEvent动画...")
    df_stopwork_animations = parse_stopwork_animations(data)

    print("正在解析中断场景信息...")
    df_interruptions = parse_interruption_scenarios(data)

    print("正在解析事件执行标签...")
    df_tags = parse_event_execution_tags(data)

    # 保存为CSV文件
    print("\n正在保存CSV文件...")
    df_summary.to_csv(output_dir / 'scene_summary.csv', index=False, encoding='utf-8-sig')
    print(f"  - scene_summary.csv (场景汇总: {len(df_summary)} 行)")

    df_nodes.to_csv(output_dir / 'scene_node_types.csv', index=False, encoding='utf-8-sig')
    print(f"  - scene_node_types.csv (节点类型: {len(df_nodes)} 行)")

    df_workspots.to_csv(output_dir / 'scene_workspots.csv', index=False, encoding='utf-8-sig')
    print(f"  - scene_workspots.csv (工作点: {len(df_workspots)} 行)")

    df_animsets.to_csv(output_dir / 'scene_animsets.csv', index=False, encoding='utf-8-sig')
    print(f"  - scene_animsets.csv (Animsets: {len(df_animsets)} 行)")

    df_animations.to_csv(output_dir / 'scene_animations.csv', index=False, encoding='utf-8-sig')
    print(f"  - scene_animations.csv (动画: {len(df_animations)} 行)")

    df_playsk_animations.to_csv(output_dir / 'scene_playsk_animations.csv', index=False, encoding='utf-8-sig')
    print(f"  - scene_playsk_animations.csv (PlaySkAnim动画: {len(df_playsk_animations)} 行)")

    df_changeidle_animations.to_csv(output_dir / 'scene_changeidle_animations.csv', index=False, encoding='utf-8-sig')
    print(f"  - scene_changeidle_animations.csv (ChangeIdleAnim动画: {len(df_changeidle_animations)} 行)")

    df_changework_animations.to_csv(output_dir / 'scene_changework_animations.csv', index=False, encoding='utf-8-sig')
    print(f"  - scene_changework_animations.csv (ChangeWorkEvent动画: {len(df_changework_animations)} 行)")

    df_stopwork_animations.to_csv(output_dir / 'scene_stopwork_animations.csv', index=False, encoding='utf-8-sig')
    print(f"  - scene_stopwork_animations.csv (StopWorkEvent动画: {len(df_stopwork_animations)} 行)")

    df_interruptions.to_csv(output_dir / 'scene_interruptions.csv', index=False, encoding='utf-8-sig')
    print(f"  - scene_interruptions.csv (中断场景: {len(df_interruptions)} 行)")

    df_tags.to_csv(output_dir / 'scene_event_tags.csv', index=False, encoding='utf-8-sig')
    print(f"  - scene_event_tags.csv (事件标签: {len(df_tags)} 行)")

    # 可选：保存为单个Excel文件的多个工作表
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        excel_file = output_dir / 'scene_analysis.xlsx'
        print(f"\n正在保存Excel文件: {excel_file}")
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # 写入各个工作表
            sheets_data = [
                (df_summary, '场景汇总'),
                (df_nodes, '节点类型'),
                (df_workspots, '工作点'),
                (df_animsets, 'Animsets'),
                (df_animations, '动画'),
                (df_playsk_animations, 'PlaySkAnim动画'),
                (df_changeidle_animations, 'ChangeIdleAnim动画'),
                (df_changework_animations, 'ChangeWorkEvent动画'),
                (df_stopwork_animations, 'StopWorkEvent动画'),
                (df_interruptions, '中断场景'),
                (df_tags, '事件标签')
            ]

            for df, sheet_name in sheets_data:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # 获取工作表对象
                worksheet = writer.sheets[sheet_name]

                # 设置表头样式
                header_font = Font(bold=True, color='FFFFFF', size=11)
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')

                # 应用表头样式
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # 设置行高
                worksheet.row_dimensions[1].height = 25  # 表头行高
                for row in range(2, worksheet.max_row + 1):
                    worksheet.row_dimensions[row].height = 20  # 数据行高

                # 自动调整列宽
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)

                    for cell in column:
                        try:
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass

                    # 设置列宽，最小12，最大60
                    adjusted_width = min(max(max_length + 2, 12), 60)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # 冻结首行
                worksheet.freeze_panes = 'A2'

                # 添加自动筛选
                worksheet.auto_filter.ref = worksheet.dimensions

                # 设置数据单元格对齐方式
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(vertical='center', wrap_text=False)

        print(f"  Excel文件保存成功! (已优化格式)")
    except ImportError:
        print("  提示: 需要安装openpyxl才能保存Excel文件")
        print("  运行: pip install openpyxl")

    print("\n解析完成!")
    print(f"\n数据统计:")
    print(f"  - 场景总数: {len(df_summary)}")
    print(f"  - 节点类型记录: {len(df_nodes)}")
    print(f"  - 工作点总数: {len(df_workspots)}")
    print(f"  - Animsets记录: {len(df_animsets)}")
    print(f"  - 动画记录: {len(df_animations)}")
    print(f"  - PlaySkAnim动画记录: {len(df_playsk_animations)}")
    print(f"  - ChangeIdleAnim动画记录: {len(df_changeidle_animations)}")
    print(f"  - ChangeWorkEvent动画记录: {len(df_changework_animations)}")
    print(f"  - StopWorkEvent动画记录: {len(df_stopwork_animations)}")
    print(f"  - 中断场景记录: {len(df_interruptions)}")
    print(f"  - 事件标签记录: {len(df_tags)}")


if __name__ == '__main__':
    main()
